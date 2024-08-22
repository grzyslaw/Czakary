import sys
import os
from openpyxl import load_workbook
import csv
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def read_excel(filename):
    logging.info(f"Reading file: {filename}")
    workbook = load_workbook(filename, data_only=True)
    data = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            row_data['Sheet'] = sheet_name
            data.append(row_data)
    logging.info(f"Read {len(data)} rows from {filename}")
    return data

def write_csv(filename, fieldnames, data):
    logging.info(f"Writing file: {filename}")
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        for row in data:
            writer.writerow({k: str(row.get(k, '')) for k in fieldnames})
    logging.info(f"Wrote {len(data)} rows to {filename}")

def create_lookup(data, name_key, container_key, filename):
    return {(row[name_key], row[container_key]): (row['Datatype'], filename, row['Sheet']) 
            for row in data 
            if all(key in row for key in [name_key, container_key, 'Datatype', 'Sheet'])}

def process_data(input_sheets, selections):
    lookups = [create_lookup(sheet, 'Name', 'Container', os.path.basename(file)) for sheet, file in input_sheets]

    updated_selections = []
    mismatches = []

    for row in selections:
        in_key = (row.get('In_Name', ''), row.get('In_Container', ''))
        out_key = (row.get('Out_Name', ''), row.get('Out_Container', ''))
        
        in_datatype_info = next(((datatype, filename, sheet) for lookup in lookups for key, (datatype, filename, sheet) in lookup.items() if key == in_key), ('Unknown', 'Unknown', 'Unknown'))
        out_datatype_info = next(((datatype, filename, sheet) for lookup in lookups for key, (datatype, filename, sheet) in lookup.items() if key == out_key), ('Unknown', 'Unknown', 'Unknown'))

        row['In_Datatype'] = in_datatype_info[0]
        row['Out_Datatype'] = out_datatype_info[0]
        row['In_File'] = in_datatype_info[1]
        row['In_Sheet'] = in_datatype_info[2]
        row['Out_File'] = out_datatype_info[1]
        row['Out_Sheet'] = out_datatype_info[2]

        updated_selections.append(row)

        if row['In_Datatype'] != row['Out_Datatype']:
            mismatches.append(row)

    logging.info(f"Processed {len(updated_selections)} selections, found {len(mismatches)} mismatches")
    return updated_selections, mismatches

def main(sheets_folder):
    try:
        # Check if the sheets folder exists
        if not os.path.exists(sheets_folder):
            logging.error(f"Error: The folder {sheets_folder} does not exist")
            sys.exit(1)

        # Get all Excel files in the sheets folder
        excel_files = [f for f in os.listdir(sheets_folder) if f.endswith('.xlsx')]
        
        if len(excel_files) < 1:
            logging.error("Error: No Excel files found in the sheets folder")
            sys.exit(1)

        # Read all input sheets
        input_sheets = [(read_excel(os.path.join(sheets_folder, file)), file) for file in excel_files]
        
        # Read selections file (assumed to be in the same directory as the script)
        script_dir = os.path.dirname(os.path.abspath(__file__))
        selections_file = os.path.join(script_dir, 'Selections.xlsx')
        if not os.path.exists(selections_file):
            logging.error(f"Error: Selections.xlsx not found in {script_dir}")
            sys.exit(1)
        selections = read_excel(selections_file)

        updated_selections, mismatches = process_data(input_sheets, selections)

        # Write updated selections
        selection_fieldnames = ['In_Name', 'In_Container', 'In_Datatype', 'In_File', 'In_Sheet', 'Out_Name', 'Out_Container', 'Out_Datatype', 'Out_File', 'Out_Sheet']
        write_csv('Updated_Selections.csv', selection_fieldnames, updated_selections)

        # Write mismatches
        write_csv('Datatype_Mismatches.csv', selection_fieldnames, mismatches)

    except Exception as e:
        logging.exception(f"An error occurred: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python dupa.py /path/to/sheets/folder")
        sys.exit(1)
    main(sys.argv[1])